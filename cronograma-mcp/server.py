import os
import re
import uuid
import base64
import logging
from time import perf_counter
from datetime import timedelta
from typing import List, Optional, Dict, Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.comments import Comment

from mcp.server.fastmcp import FastMCP


# -----------------------------
# Versionamento de formato
# -----------------------------
FORMAT_VERSION = "1.0.1"


# -----------------------------
# Config (conservador e previsível)
# -----------------------------
DEFAULT_OUTPUT_DIR = os.getenv("CRONOGRAMA_OUTPUT_DIR", os.path.abspath("./outputs"))
MAX_ROWS = int(os.getenv("CRONOGRAMA_MAX_ROWS", "300"))

os.makedirs(DEFAULT_OUTPUT_DIR, exist_ok=True)

mcp = FastMCP("Cronograma-MCP", json_response=True)


# -----------------------------
# Logging (estruturado e útil)
# -----------------------------
LOG_LEVEL = os.getenv("CRONOGRAMA_LOG_LEVEL", "INFO").upper()
logging.basicConfig(
    level=getattr(logging, LOG_LEVEL, logging.INFO),
    format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
)
logger = logging.getLogger("mcp-cronograma")


# -----------------------------
# Helpers
# -----------------------------
_TIME_RE = re.compile(r"^\d{1,5}:\d{2}:\d{2}$")  # até 99999h

EXCEL_CELL_CHAR_LIMIT = 32767


def _sanitize_text(text: str, max_len: int = EXCEL_CELL_CHAR_LIMIT) -> str:
    """
    Evita problemas clássicos:
    - Excel tem limite de 32767 chars por célula
    - remove caracteres de controle que podem quebrar arquivos
    """
    if text is None:
        return ""
    if not isinstance(text, str):
        text = str(text)

    # Remove caracteres de controle (exceto \n \t)
    text = re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", text)

    if len(text) > max_len:
        text = text[:max_len]
    return text


def _parse_hhmmss(value: str) -> timedelta:
    """
    Aceita 'H:MM:SS' com horas grandes.
    """
    value = (value or "").strip()
    if not _TIME_RE.match(value):
        raise ValueError(f"Duração inválida: '{value}'. Use formato H:MM:SS (ex.: 2:00:00).")

    h_str, m_str, s_str = value.split(":")
    h = int(h_str)
    m = int(m_str)
    s = int(s_str)
    if m > 59 or s > 59:
        raise ValueError(f"Duração inválida: '{value}'. Minutos/Segundos devem ser <= 59.")

    return timedelta(hours=h, minutes=m, seconds=s)


def _safe_filename(name: str) -> str:
    name = _sanitize_text(name or "cronograma").strip()
    name = re.sub(r"[^a-zA-Z0-9._-]+", "_", name)
    return name[:120] if len(name) > 120 else name


def _build_xlsx(model_rows: List[Dict[str, Any]], output_path: str, titulo: Optional[str], request_id: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Cronograma"

    headers = ["Nome da Tarefa", "Duração", "Responsável"]
    ws.append(headers)

    # Estilos
    header_fill = PatternFill("solid", fgColor="D9D9D9")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Cabeçalho
    for col in range(1, 4):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 28

    # Metadados (comentário no cabeçalho)
    meta_comment = f"Projeto: {titulo or '-'}\nFormato: {FORMAT_VERSION}\nRequestId: {request_id}"
    ws["A1"].comment = Comment(meta_comment, "MCP")

    # Linhas
    for row in model_rows:
        nome = _sanitize_text(row.get("nome_tarefa", "")).strip()
        dur = _sanitize_text(row.get("duracao_hhmmss", "")).strip()
        resp = _sanitize_text(row.get("responsavel", "")).strip()

        td = _parse_hhmmss(dur) if dur else None
        ws.append([nome, td, resp])

    # Formatação de duração
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=2).number_format = "[h]:mm:ss"
        for c in range(1, 4):
            ws.cell(row=r, column=c).border = border
            ws.cell(row=r, column=c).alignment = left if c != 2 else center

    # Comentários do modelo (guia in-line)
    if ws.max_row >= 2:
        ws["B2"].comment = Comment("HORAS TOTAIS DA MACRO ATIVIDADE em HORAS. Ex: 10:00:00", "MCP")
    if ws.max_row >= 3:
        ws["B3"].comment = Comment("Horas da micro atividade. Ex: 02:00:00", "MCP")

    ws.freeze_panes = "A2"

    # Tabela Excel para filtros
    last_row = max(ws.max_row, 2)
    table_ref = f"A1:C{last_row}"
    tab = Table(displayName="TabelaCronograma", ref=table_ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)

    # Aba de instruções
    ws2 = wb.create_sheet("MCP_Instruções")
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 90
    ws2.append(["Campo", "Como usar"])

    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    rows = [
        ("Objetivo", "Padronizar a confecção do cronograma do projeto em formato simples, repetível e auditável."),
        ("Modelo", "Nome da Tarefa | Duração | Responsável (Macro atividade e Micros atividades)."),
        ("Duração", "Use H:MM:SS (ex.: 10:00:00). A coluna é formatada como [h]:mm:ss."),
        ("Responsável", "Informe o papel (ex.: Arquiteto de Solução)."),
        ("Versão do Formato", FORMAT_VERSION),
        ("Checklist", "1) Macro definida  2) Micros listadas  3) Durações  4) Responsáveis  5) Revisão final"),
    ]

    for k, v in rows:
        ws2.append([_sanitize_text(k), _sanitize_text(v)])

    for r in range(2, 2 + len(rows)):
        for c in range(1, 3):
            ws2.cell(row=r, column=c).alignment = Alignment(vertical="top", wrap_text=True)
            ws2.cell(row=r, column=c).border = border

    ws2.freeze_panes = "A2"

    wb.save(output_path)


# -----------------------------
# MCP tools
# -----------------------------
@mcp.tool()
def health_check() -> Dict[str, Any]:
    """
    Health check do MCP: valida diretório de output e limites.
    Ideal para o n8n chamar antes do fluxo.
    """
    writable = os.access(DEFAULT_OUTPUT_DIR, os.W_OK)
    return {
        "status": "healthy" if writable else "degraded",
        "output_dir": DEFAULT_OUTPUT_DIR,
        "writable": writable,
        "max_rows": MAX_ROWS,
        "format_version": FORMAT_VERSION,
    }


@mcp.tool()
def gerar_cronograma_xlsx(
    projeto: str,
    linhas: List[Dict[str, str]],
    nome_arquivo: Optional[str] = None,
    retornar_base64: bool = True,
) -> Dict[str, Any]:
    """
    Gera um arquivo XLSX no modelo:
    Nome da Tarefa | Duração | Responsável

    Cada item em 'linhas' deve conter:
      - nome_tarefa (str) obrigatório
      - duracao_hhmmss (str) opcional, formato H:MM:SS
      - responsavel (str) opcional

    Retorno:
      - output_path
      - filename
      - base64 (opcional)
      - métricas simples
      - format_version
      - request_id
    """
    request_id = str(uuid.uuid4())
    t0 = perf_counter()

    if not projeto or not isinstance(projeto, str):
        raise ValueError("Campo 'projeto' é obrigatório e deve ser string.")

    if not isinstance(linhas, list) or len(linhas) == 0:
        raise ValueError("Campo 'linhas' deve ser uma lista com pelo menos 1 item.")

    if len(linhas) > MAX_ROWS:
        raise ValueError(f"Muitas linhas ({len(linhas)}). Limite atual: {MAX_ROWS}.")

    if not os.access(DEFAULT_OUTPUT_DIR, os.W_OK):
        raise RuntimeError(f"Diretório de output não é gravável: {DEFAULT_OUTPUT_DIR}")

    logger.info(
        f"[{request_id}] iniciar_geracao projeto='{projeto}' linhas={len(linhas)} retornar_base64={retornar_base64}"
    )

    normalized: List[Dict[str, str]] = []
    for i, row in enumerate(linhas, start=1):
        nome = _sanitize_text(row.get("nome_tarefa", "")).strip()
        dur = _sanitize_text(row.get("duracao_hhmmss", "")).strip()
        resp = _sanitize_text(row.get("responsavel", "")).strip()

        if not nome:
            raise ValueError(f"Linha {i}: 'nome_tarefa' é obrigatório.")

        if dur:
            _parse_hhmmss(dur)  # valida formato

        normalized.append({"nome_tarefa": nome, "duracao_hhmmss": dur, "responsavel": resp})

    base_name = _safe_filename(nome_arquivo or f"cronograma_{projeto}")
    if not base_name.lower().endswith(".xlsx"):
        base_name += ".xlsx"

    output_path = os.path.join(DEFAULT_OUTPUT_DIR, base_name)

    _build_xlsx(normalized, output_path=output_path, titulo=_sanitize_text(projeto), request_id=request_id)

    file_size = os.path.getsize(output_path)
    elapsed_ms = int((perf_counter() - t0) * 1000)

    result: Dict[str, Any] = {
        "request_id": request_id,
        "format_version": FORMAT_VERSION,
        "filename": base_name,
        "output_path": output_path,
        "rows": len(normalized),
        "output_dir": DEFAULT_OUTPUT_DIR,
        "file_size_bytes": file_size,
        "elapsed_ms": elapsed_ms,
    }

    if retornar_base64:
        with open(output_path, "rb") as f:
            result["base64"] = base64.b64encode(f.read()).decode("utf-8")

        result["content_type"] = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        result["download_name"] = os.path.basename(output_path)


    logger.info(
        f"[{request_id}] geracao_concluida arquivo='{base_name}' size_bytes={file_size} elapsed_ms={elapsed_ms}"
    )

    return result


if __name__ == "__main__":
    mcp.run()

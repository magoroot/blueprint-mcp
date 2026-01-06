#!/usr/bin/env python3
"""
Servidor de Cronograma (produção)
- HTTP-first (FastAPI + Uvicorn) para uso em Docker/OpenWebUI
- Download via /download/{token}
- Endpoints /cronograma/generate e /cronograma/validate
- Modo MCP stdio opcional via CRONOGRAMA_RUN_MODE=stdio (para execução por cliente MCP)
"""

import os
import sys
import re
import unicodedata
import logging
import base64
import secrets
import asyncio
import threading
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, Any, Optional, Tuple, List

from fastapi import FastAPI, HTTPException
from fastapi.responses import FileResponse, JSONResponse
import uvicorn

# MCP (opcional)
from mcp.server.fastmcp import FastMCP

# XLSX
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Pydantic (essencial pro OpenAPI "travar" requestBody)
from pydantic import BaseModel, Field, ConfigDict
from typing_extensions import Annotated


# ========================================================
# CONFIG
# ========================================================

LOG_LEVEL = os.getenv("LOG_LEVEL", "INFO").upper()
logging.basicConfig(
    level=getattr(logging, LOG_LEVEL, logging.INFO),
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger("cronograma-mcp")

OUTPUT_DIR = Path(os.getenv("CRONOGRAMA_OUTPUT_DIR", "./outputs"))
MAX_ROWS = int(os.getenv("CRONOGRAMA_MAX_ROWS", "500"))
TTL_MINUTES = int(os.getenv("CRONOGRAMA_TTL_MINUTES", "30"))
BASE_URL = os.getenv("CRONOGRAMA_BASE_URL", "http://localhost:8000").rstrip("/")
HTTP_PORT = int(os.getenv("CRONOGRAMA_HTTP_PORT", "8000"))

# http (padrão) | stdio (apenas quando cliente MCP spawnar)
RUN_MODE = os.getenv("CRONOGRAMA_RUN_MODE", "http").strip().lower()

# Cleanup periódico (segundos)
CLEANUP_INTERVAL_SECONDS = int(os.getenv("CRONOGRAMA_CLEANUP_INTERVAL_SECONDS", "60"))

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Registry: token -> {filepath, filename, expires_at}
file_registry: Dict[str, Dict[str, Any]] = {}
registry_lock = threading.Lock()


# ========================================================
# Pydantic Models (OpenAPI rígido)
# ========================================================

class ProjectModel(BaseModel):
    model_config = ConfigDict(extra="allow")
    name: str = Field(..., min_length=1, description="Nome do projeto")
    owner: Optional[str] = Field(default=None, description="Responsável/owner do projeto")

class MicroModel(BaseModel):
    model_config = ConfigDict(extra="allow")
    name: str = Field(..., min_length=1, description="Nome da micro atividade")
    hours: float = Field(..., gt=0, description="Horas da micro atividade (deve ser > 0)")
    responsible: Optional[str] = Field(default=None, description="Responsável pela micro")

class MacroModel(BaseModel):
    model_config = ConfigDict(extra="allow")
    name: str = Field(..., min_length=1, description="Nome da macro atividade")
    responsible: Optional[str] = Field(default=None, description="Responsável pela macro")
    micros: List[MicroModel] = Field(..., min_length=1, description="Lista de micros (obrigatório, >= 1)")

class SettingsModel(BaseModel):
    model_config = ConfigDict(extra="allow")
    format_version: Optional[str] = Field(default="1.0.0")
    sheet_name: Optional[str] = Field(default="Planilha1")
    include_project_row: Optional[bool] = Field(default=True)
    max_rows: Optional[int] = Field(default=None, ge=1, description="Override do limite de linhas")

class PayloadModel(BaseModel):
    model_config = ConfigDict(extra="allow")
    project: ProjectModel
    macros: List[MacroModel] = Field(..., min_length=1)
    settings: Optional[SettingsModel] = Field(default=None)


# ========================================================
# UTIL
# ========================================================

def sanitize_filename(name: str) -> str:
    """Gera nome de arquivo seguro (ASCII) e previsível.

    - Normaliza Unicode (remove acentos)
    - Remove caracteres inválidos / controles
    - Troca espaços por underscore
    - Remove tudo que não for [A-Za-z0-9._-]
    - Limita tamanho para evitar nomes gigantes
    """
    if not name:
        return "Cronograma"

    # Normaliza e remove acentos (vira ASCII)
    n = unicodedata.normalize("NFKD", str(name))
    n = n.encode("ascii", "ignore").decode("ascii")

    # Troca separadores comuns por espaço (antes de virar underscore)
    n = n.replace("->", " ").replace("=>", " ").replace("/", " ")
    n = n.replace("-", " ").replace("—", " ").replace("–", " ")

    # Remove caracteres inválidos do Windows/headers e controles
    n = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "", n)

    # Espaços -> underscore e limpa
    n = re.sub(r"\s+", "_", n).strip("_")

    # Mantém somente caracteres seguros
    n = re.sub(r"[^A-Za-z0-9._-]", "", n)

    # Evita múltiplos underscores
    n = re.sub(r"_+", "_", n)

    return n[:180] if n else "Cronograma"


def hours_to_duration_display(hours: float) -> str:
    """
    Converte horas decimais para HHH:MM:SS (sem virar dias).
    Ex: 247.6667 -> 247:40:00
    """
    if hours is None or hours < 0:
        hours = 0.0
    total_seconds = int(round(hours * 3600))
    h = total_seconds // 3600
    m = (total_seconds % 3600) // 60
    s = total_seconds % 60
    return f"{h:d}:{m:02d}:{s:02d}"

def generate_token() -> str:
    return secrets.token_urlsafe(32)

def cleanup_expired_files() -> None:
    now = datetime.now()
    expired_tokens: List[str] = []

    with registry_lock:
        for token, info in list(file_registry.items()):
            expires_at = info.get("expires_at")
            if expires_at and now > expires_at:
                expired_tokens.append(token)

        for token in expired_tokens:
            info = file_registry.pop(token, None)
            if not info:
                continue
            try:
                filepath = Path(info["filepath"])
                if filepath.exists():
                    filepath.unlink()
                    logger.info(f"Arquivo expirado removido: {filepath.name}")
            except Exception as e:
                logger.error(f"Erro ao remover arquivo expirado: {e}")


# ========================================================
# VALIDATION (regras adicionais além do Pydantic)
# ========================================================

def validate_payload_dict(payload: dict) -> Tuple[bool, Optional[dict]]:
    """
    Mantém validações "de negócio" extras + MAX_ROWS.
    Observação: Pydantic já garante required + tipos + min_length + hours>0.
    """
    errors = []

    # MAX_ROWS
    settings = payload.get("settings") if isinstance(payload.get("settings"), dict) else {}
    try:
        max_rows_limit = int(settings.get("max_rows", MAX_ROWS))
    except Exception:
        max_rows_limit = MAX_ROWS

    total_rows = 1  # projeto
    macros = payload.get("macros", [])
    for macro in macros:
        total_rows += 1  # macro
        micros = (macro or {}).get("micros", [])
        total_rows += len(micros)

    if total_rows > max_rows_limit:
        errors.append({
            "field": "total_rows",
            "issue": f"total de linhas ({total_rows}) excede o limite ({max_rows_limit})"
        })
        return False, {
            "ok": False,
            "error_code": "MAX_ROWS_EXCEEDED",
            "message": f"O cronograma possui {total_rows} linhas, excedendo o limite de {max_rows_limit}",
            "details": errors,
        }

    return True, None


# ========================================================
# XLSX GEN
# ========================================================

def generate_xlsx(payload: dict) -> Tuple[Path, dict, float]:
    project = payload["project"]
    settings = payload.get("settings", {}) if isinstance(payload.get("settings", {}), dict) else {}
    macros = payload["macros"]

    sheet_name = settings.get("sheet_name", "Planilha1")
    include_project_row = bool(settings.get("include_project_row", True))

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # styles
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    project_font = Font(bold=True, size=11)
    project_fill = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")

    macro_font = Font(bold=True, size=10)
    macro_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")

    border_thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    ws.column_dimensions["A"].width = 70
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 22

    # header row
    ws["A1"] = "Nome da Tarefa"
    ws["B1"] = "Duration"
    ws["C1"] = "Responsável"

    for col in ["A", "B", "C"]:
        cell = ws[f"{col}1"]
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border_thin
        cell.alignment = header_alignment if col != "A" else Alignment(horizontal="left", vertical="center")

    ws.freeze_panes = "A2"
    current_row = 2

    # totals
    project_total_hours = 0.0
    macro_summaries = []

    for macro in macros:
        macro_total_hours = 0.0
        micro_count = 0
        for micro in macro["micros"]:
            h = float(micro["hours"])
            macro_total_hours += h
            micro_count += 1

        project_total_hours += macro_total_hours
        macro_summaries.append({
            "name": macro["name"],
            "hours": round(macro_total_hours, 4),
            "duration_display": hours_to_duration_display(macro_total_hours),
            "micro_count": micro_count,
        })

    # project row
    if include_project_row:
        ws[f"A{current_row}"] = project["name"]
        ws[f"B{current_row}"] = hours_to_duration_display(project_total_hours)  # texto
        ws[f"C{current_row}"] = project.get("owner", "")

        for col in ["A", "B", "C"]:
            cell = ws[f"{col}{current_row}"]
            cell.font = project_font
            cell.fill = project_fill
            cell.border = border_thin
            if col == "B":
                cell.alignment = Alignment(horizontal="center", vertical="center")
        current_row += 1

    # macros + micros
    for macro_idx, macro in enumerate(macros):
        ws[f"A{current_row}"] = macro["name"]
        ws[f"B{current_row}"] = macro_summaries[macro_idx]["duration_display"]  # texto
        ws[f"C{current_row}"] = macro.get("responsible", "")

        for col in ["A", "B", "C"]:
            cell = ws[f"{col}{current_row}"]
            cell.font = macro_font
            cell.fill = macro_fill
            cell.border = border_thin
            if col == "B":
                cell.alignment = Alignment(horizontal="center", vertical="center")

        current_row += 1

        for micro in macro["micros"]:
            hours = float(micro["hours"])
            ws[f"A{current_row}"] = f"    {micro['name']}"
            ws[f"B{current_row}"] = hours_to_duration_display(hours)  # texto
            ws[f"C{current_row}"] = micro.get("responsible", "")

            for col in ["A", "B", "C"]:
                cell = ws[f"{col}{current_row}"]
                cell.border = border_thin
                if col == "B":
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            current_row += 1

    project_name_clean = sanitize_filename(project["name"])
    timestamp = datetime.now().strftime("%Y-%m-%d")
    filename = f"Cronograma_-_{project_name_clean}_-_{timestamp}.xlsx"
    filepath = OUTPUT_DIR / filename

    wb.save(filepath)
    logger.info(f"XLSX gerado: {filepath.name}")

    summary = {
        "macro_count": len(macros),
        "micro_count": sum(m["micro_count"] for m in macro_summaries),
        "macros": macro_summaries,
    }

    return filepath, summary, project_total_hours


# ========================================================
# CORE SERVICE (compartilhado)
# ========================================================

def build_generation_response(payload: dict) -> dict:
    logger.info("Iniciando geração de cronograma XLSX")

    cleanup_expired_files()

    # validação extra (MAX_ROWS, etc.)
    ok, err = validate_payload_dict(payload)
    if not ok:
        logger.warning(f"Validação falhou: {err['error_code']}")
        return err

    filepath, summary, project_total_hours = generate_xlsx(payload)

    with open(filepath, "rb") as f:
        file_base64 = base64.b64encode(f.read()).decode("utf-8")

    token = generate_token()
    expires_at = datetime.now() + timedelta(minutes=TTL_MINUTES)

    with registry_lock:
        file_registry[token] = {
            "filepath": str(filepath),
            "filename": filepath.name,
            "expires_at": expires_at,
        }

    download_url = f"{BASE_URL}/download/{token}"

    project = payload["project"]
    settings = payload.get("settings", {}) if isinstance(payload.get("settings", {}), dict) else {}

    resp = {
        "ok": True,
        "format_version": settings.get("format_version", "1.0.0"),
        "project_name": project["name"],
        "project_total_hours": round(project_total_hours, 4),
        "project_total_duration_display": hours_to_duration_display(project_total_hours),
        "filename": filepath.name,
        "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "base64": file_base64,
        "download_url": download_url,
        "download_expires_at": expires_at.isoformat(),
        "summary": summary,
    }

    logger.info(f"Cronograma gerado com sucesso: {filepath.name}")
    logger.info(f"Download disponível em: {download_url}")
    return resp


def build_validate_response(payload: dict) -> dict:
    logger.info("Validando payload (sem gerar arquivo)")

    ok, err = validate_payload_dict(payload)
    if not ok:
        return err

    project_total_hours = 0.0
    macro_count = len(payload["macros"])
    micro_count = 0

    for macro in payload["macros"]:
        for micro in macro["micros"]:
            project_total_hours += float(micro["hours"])
            micro_count += 1

    return {
        "ok": True,
        "message": "Payload válido",
        "preview": {
            "project_name": payload["project"]["name"],
            "project_total_hours": round(project_total_hours, 4),
            "project_total_duration_display": hours_to_duration_display(project_total_hours),
            "macro_count": macro_count,
            "micro_count": micro_count,
        },
    }


# ========================================================
# MCP (opcional)
# ========================================================

mcp = FastMCP("cronograma-mcp")

@mcp.tool()
def gerar_xlsx(payload: dict) -> dict:
    return build_generation_response(payload)

@mcp.tool()
def validar(payload: dict) -> dict:
    return build_validate_response(payload)

@mcp.tool()
def health() -> dict:
    with registry_lock:
        active_files = len(file_registry)
    return {
        "ok": True,
        "service": "cronograma-mcp",
        "status": "healthy",
        "output_dir": str(OUTPUT_DIR),
        "max_rows": MAX_ROWS,
        "ttl_minutes": TTL_MINUTES,
        "active_files": active_files,
        "run_mode": RUN_MODE,
    }


# ========================================================
# HTTP (produção)
# ========================================================

app = FastAPI(title="Cronograma Server (HTTP-first)")

@app.get("/health")
async def http_health():
    return {
        "ok": True,
        "service": "cronograma-http",
        "status": "healthy",
        "run_mode": RUN_MODE,
    }

@app.post("/cronograma/generate")
async def http_generate(payload: PayloadModel):
    """
    PayloadModel deixa o OpenAPI rígido:
    - requestBody REQUIRED
    - project REQUIRED
    - macros REQUIRED (>=1)
    - micros REQUIRED (>=1)
    - hours > 0
    """
    payload_dict = payload.model_dump(exclude_none=True)
    resp = build_generation_response(payload_dict)
    status = 200 if resp.get("ok") else 400
    return JSONResponse(status_code=status, content=resp)

@app.post("/cronograma/validate")
async def http_validate(payload: PayloadModel):
    payload_dict = payload.model_dump(exclude_none=True)
    resp = build_validate_response(payload_dict)
    status = 200 if resp.get("ok") else 400
    return JSONResponse(status_code=status, content=resp)

@app.get("/download/{token}")
async def download_file(token: str):
    cleanup_expired_files()

    with registry_lock:
        info = file_registry.get(token)

    if not info:
        raise HTTPException(status_code=404, detail="Arquivo não encontrado ou expirado")

    filepath = Path(info["filepath"])
    if not filepath.exists():
        with registry_lock:
            file_registry.pop(token, None)
        raise HTTP는Exception(status_code=404, detail="Arquivo não encontrado")

    logger.info(f"Download iniciado: {info['filename']}")
    return FileResponse(
        path=str(filepath),
        filename=info["filename"],
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{info["filename"]}"'},
    )

@app.on_event("startup")
async def on_startup():
    logger.info("HTTP server startup - iniciando tarefa de cleanup periódico")

    async def _periodic_cleanup():
        while True:
            try:
                cleanup_expired_files()
            except Exception as e:
                logger.error(f"Erro no cleanup periódico: {e}")
            await asyncio.sleep(CLEANUP_INTERVAL_SECONDS)

    asyncio.create_task(_periodic_cleanup())


# ========================================================
# MAIN
# ========================================================

async def run_mcp_stdio():
    """
    Modo MCP stdio: SOMENTE quando um cliente MCP spawnar este processo.
    Observação: versões do fastmcp diferem. Vamos tentar o modo mais compatível.
    """
    logger.info("MCP stdio iniciado")

    # Caminho 1 (comum em FastMCP): mcp.run(transport="stdio")
    try:
        await mcp.run(transport="stdio")
        return
    except TypeError:
        pass
    except Exception as e:
        logger.warning(f"Falha ao iniciar MCP com transport='stdio': {e}")

    # Caminho 2: fallback se sua lib suportar stdio_server + run(streams)
    try:
        from mcp.server.stdio import stdio_server
        async with stdio_server() as (read_stream, write_stream):
            await mcp.run(read_stream, write_stream)  # algumas versões aceitam
            return
    except Exception as e:
        logger.error(f"Erro fatal no MCP stdio: {e}", exc_info=True)
        raise

def run_http():
    logger.info(f"Iniciando HTTP na porta {HTTP_PORT}")
    uvicorn.run(app, host="0.0.0.0", port=HTTP_PORT, log_level="info")

if __name__ == "__main__":
    logger.info("=" * 60)
    logger.info("Cronograma Server")
    logger.info("=" * 60)
    logger.info(f"RUN_MODE: {RUN_MODE}")
    logger.info(f"OUTPUT_DIR: {OUTPUT_DIR}")
    logger.info(f"MAX_ROWS: {MAX_ROWS}")
    logger.info(f"TTL_MINUTES: {TTL_MINUTES}")
    logger.info(f"BASE_URL: {BASE_URL}")
    logger.info(f"HTTP_PORT: {HTTP_PORT}")
    logger.info("=" * 60)

    if RUN_MODE == "stdio":
        try:
            asyncio.run(run_mcp_stdio())
        except KeyboardInterrupt:
            logger.info("Encerrado pelo usuário")
        except Exception:
            sys.exit(1)
    else:
        run_http()
